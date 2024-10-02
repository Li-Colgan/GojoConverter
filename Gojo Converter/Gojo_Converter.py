'''
GOJO CONVERTER
Converts Menus in Gojo's 2023 xlsx format to Square Bulk Upload format.
'''
# libs
import openpyxl
import re

# CONSTANTS
source_wb = openpyxl.load_workbook(r'C:\Users\OWR Laptop\Desktop\Gojo Converter\Gojo_Menu.xlsx')
target_wb = openpyxl.load_workbook(r'C:\Users\OWR Laptop\Desktop\Gojo Converter\Square_Bulk_Upload_Template.xlsx')
target_sheet = target_wb.active
rows_to_search = 100  # top-down rows to search for headers
reset_start_row = 3  # deletes rows from this down inclusive for init

# Define headers and their corresponding target columns
headers_and_locations = {
    'Item Description (as advertised on your menu board)': 7,  # G column
    'POS Description': 3,
    'Category': 8,
    'Price*': 20,
    'Weight**': 15,
}

# Clear all values from row 3 down in the target sheet
for row in target_sheet.iter_rows(min_row=reset_start_row):
    for cell in row:
        cell.value = None

# FUNCTIONS
# removes non-numeric characters and converts to kg or litres
def convert_weight(weight_str):
    try:
        weight_str = re.sub(r'[^0-9.]', '', weight_str).strip()  
        if weight_str:  
            weight_grams = float(weight_str)  
            return weight_grams / 1000  
        else:
            print(f"No valid numeric value found in weight '{weight_str}'. Skipping.")
            return None  
    except ValueError:
        print(f"Invalid weight value '{weight_str}' for conversion. Skipping.")
        return None  

# TRANSFER LOOP
# sheet loop
for sheet_name in source_wb.sheetnames:
    source_sheet = source_wb[sheet_name] 
    print(f"Processing sheet: {sheet_name}")

    # Loop through all headers for each sheet
    for header, target_col in headers_and_locations.items():
        source_col = None
        
        # Find header
        for row in source_sheet.iter_rows(max_row=rows_to_search):  
            for cell in row:
                if cell.value == header:
                    source_col = cell.column
                    header_row = cell.row  
                    break
            if source_col:
                break
        
        # Debug message
        if not source_col:
            print(f"Header '{header}' not found in the sheet '{sheet_name}'.")
            continue

        # Find the next empty row in the target sheet
        target_row = reset_start_row
        while target_sheet.cell(row=target_row, column=target_col).value is not None:
            target_row += 1

        empty_counter = 0  # Counter for empty cells

        # Header loop
        for row in source_sheet.iter_rows(min_row=header_row + 1, min_col=source_col, max_col=source_col):
            # Account for merged cells
            cell_value = row[0].value if not isinstance(row[0], openpyxl.cell.MergedCell) else source_sheet.cell(row=row[0].row, column=source_col).value
            
            if cell_value:  
                empty_counter = 0  # Reset counter since we found a non-empty cell
                
                # Convert weight
                if header == 'Weight**':
                    if isinstance(cell_value, (str, int, float)):
                        weight_kg = convert_weight(cell_value)
                        if weight_kg is not None:
                            target_sheet.cell(row=target_row, column=target_col, value=weight_kg)
                else:
                    target_sheet.cell(row=target_row, column=target_col, value=cell_value)
                
                target_row += 1  # Move to the next target row
            else:
                empty_counter += 1  # Increment counter if the cell is empty
                target_row += 1
            
            # Check if we've encountered 20 consecutive empty cells
            if empty_counter >= 20:
                print(f"Stopping copy for header '{header}' after 20 empty cells.")
                break  # Exit the loop if 20 empty cells are found

# Save target workbook and print success message
target_wb.save(r'C:\Users\OWR Laptop\Desktop\Gojo Converter\Square_Bulk_Upload_Template.xlsx')
print("Transfer complete.")
