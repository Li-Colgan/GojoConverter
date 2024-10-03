'''
TODO
-online visibility, delivery hardcoded, auto add to bill, item type currently hardcoded, would be good to make them configurable
-test on oneworld square acc
-likely very unforgiving if square location names don't exactly match sheet names
-account for duplicate locations (option for enabling for all or giving name for a sheet directly?)
'''

#libs
import openpyxl 
from utilities import *
from copier import *

'''
GOJO CONVERTER
Converts Menus in Gojo's 2023 xlsx format to Square Bulk Upload format.
'''

#CONFIG======================================================================================================================
source_wb = openpyxl.load_workbook(r'Gojo Converter/Gojo_Menu.xlsx') #load gojo menu
target_wb = openpyxl.load_workbook(r'Gojo Converter/Square_Bulk_Upload_Template.xlsx')# load square template
target_sheet = target_wb.worksheets[0]  #use first sheet of square template
rows_to_search = 100  #top-down rows to search for headers
reset_start_row = 3  #deletes rows from this down inclusive for init
overwrite = 1 #set to 1 to overwrite contents of square template, set to 0 to add to preexisting items

#gojo header and corresponding column in square template
headers_and_locations = {
    'Item Description (as advertised on your menu board)': 7, #Description
    'POS Description': 3, #Item Name
    'Category': 8, #Reporting Category
    'Price*': 20, #Weight (kg)
    'Weight**': 15, #Price
}
#TRANSFER LOOP==================================================================================================================
#clear template if overwrite is set
init(target_sheet, reset_start_row, overwrite)
#sheet loop
for sheet_name in source_wb.sheetnames:
    source_sheet = source_wb[sheet_name] 
    print(f"Processing sheet: {sheet_name}") #debug
    next_empty_row = find_next_empty_row(target_sheet, reset_start_row) #copy to next empty row for each sheet

    enable_column = find_enable_column(target_sheet, source_sheet.title, rows_to_search)
    disable_columns = find_disable_columns(target_sheet, source_sheet.title, rows_to_search)

    #header loop
    for header, target_column in headers_and_locations.items():
        print(f"Copying '{header}' to column '{target_column}' starting at row {next_empty_row}") #debug
        copier(source_sheet, target_sheet, header, target_column, next_empty_row, rows_to_search, enable_column, disable_columns) #copy function from copier.py

#save and report success
target_wb.save(r'Gojo Converter/Square_Bulk_Upload_Template.xlsx')
print("Transfer complete.")


