from utilities import find_header, convert_weight
import openpyxl
def copier(source_sheet, target_sheet, header_name, target_column, start_row, rows_to_search):
    #get header col and row
    header_column, header_row = find_header(source_sheet, rows_to_search, header_name)

    #copy all cells under the header in the source column to the target column
    source_row = header_row + 1  

    #range of only header column, row of 1 cell
    for row in source_sheet.iter_rows(min_row=source_row, min_col=header_column, max_col=header_column):
        cell = row[0]  #get cell

        #handle merged cells
        if isinstance(cell, openpyxl.cell.MergedCell):
            value = source_sheet.cell(row=cell.row, column=header_column).value
        else:
            value = cell.value
        #convert if weight
        if header_name == 'Weight**':
            value = convert_weight(value)

        #write value
        if value is not None:
            target_row = start_row
            while target_sheet.cell(row=target_row, column=target_column).value is not None:
                target_row += 1
            target_sheet.cell(row=target_row, column=target_column, value=value)
            start_row += 1  